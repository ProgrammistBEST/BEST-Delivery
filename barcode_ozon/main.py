from openpyxl import load_workbook
import requests
import math
import FreeSimpleGUI as sg
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.units import mm
from reportlab.graphics.barcode import code128
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import shutil
import sys
import os

# Определяем путь к ресурсам в зависимости от среды выполнения
if getattr(sys, 'frozen', False):
    # Если запущено из .exe
    base_path = sys._MEIPASS
else:
    # Если запущено как обычный .py файл
    base_path = os.path.dirname(os.path.abspath(__file__))

# Переключаем рабочую директорию
os.chdir(base_path)

pdfmetrics.registerFont(
    TTFont("Calibribd", "./fonts/calibri_bold.ttf")
)
HEADERS = {"ARMBEST": {"Client-Id": "507132", "Api-Key": "token"}, "BEST 26": {"Client-Id": "2018536", "Api-Key": "token"}}
NUM_TO_COL = {1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I", 10: "J", 11: "K", 12: "L", 13: "M", 14: "N", 15: "O", 16: "P", 17: "Q"}
READY_PATH = os.path.join(os.getcwd(), "OZON внутренние")
def filter_color(text):
    text = text.replace("бордовые", "brown")
    text = text.replace("бордовая", "brown")
    text = text.replace("бордовый", "brown")
    text = text.replace("бордо", "brown")
    text = text.replace("борд", "brown")
    text = text.replace("бор", "brown")
    text = text.replace("синий", "blue")
    text = text.replace("синие", "blue")
    text = text.replace("синяя", "blue")
    text = text.replace("син", "blue")
    text = text.replace("фиолетовый", "fio")
    text = text.replace("фиолетовые", "fio")
    text = text.replace("фиолетовая", "fio")
    text = text.replace("фиолетов", "fio")
    text = text.replace("фиолето", "fio")
    text = text.replace("фиолет", "fio")
    text = text.replace("фиоле", "fio")
    text = text.replace("фиол", "fio")
    text = text.replace("фио", "fio")
    text = text.replace("фи", "fio")
    text = text.replace("ф", "fio")
    text = text.replace("с", "blue")
    text = text.replace("б", "brown")
    return text

def get_all_data(headers):
    not_first_loop = False
    data = []
    while True:
        if not_first_loop:
            json = {"last_id": lastid}
        else:
            json = {}
        this_data = requests.post(
            "https://api-seller.ozon.ru/v2/product/list",
            headers=headers,
            json=json,
        ).json()
        print(this_data["result"]["total"])
        data += this_data["result"]["items"]
        lastid = this_data["result"]["last_id"]
        if not this_data["result"]["last_id"]:
            break
        not_first_loop = True
        if this_data["result"]["total"] < 1000:
            break
    return data

def width(self):
    self._calculate()
    return self._width

# wb = load_workbook(input("Эксель: "))
def add_pdf_page(c, brand, barcode, number=""):
    c.rotate(90)
    c.setFillColorRGB(0, 0, 0, 1)
    c.setFont("Calibribd", 28)
    c.drawCentredString(60*mm, -10*mm, "OZON")
    c.setFont("Calibribd", 18)

    if brand == "Armbest":
        c.rect(10 * mm, -13 * mm, 100 * mm, -15 * mm, 1, 1)
        c.setFillColorRGB(1, 1, 1, 1)

    c.drawCentredString(60*mm, -20*mm, brand)
    c.drawCentredString(60*mm, -26*mm, barcode)
    c.setFillColorRGB(0, 0, 0, 1)
    # if number:
    #     c.drawRightString(110*mm, -10*mm, "№ "+number)
    code_width = 121
    bar_width = 1
    while code_width > 120:
        code_im = code128.Code128(barcode, barWidth=bar_width*mm, barHeight=40*mm)
        code_width = width(code_im)/mm
        bar_width -= 0.1
    code_im.drawOn(c, ((120 * mm)/2)-(width(code_im)/2) , -70 * mm)
    c.showPage()

def excel_to_json(wb_path):
    wb = load_workbook(wb_path, data_only=True)
    ws = wb.active
    sizes = {4:"41", 5:"42", 6:"43", 7:"44", 8:"45", 9:"46", 10:"47", 11:"48", 12:None, 13:None, 14:None, 15:None}
    size_row = 4
    models = {}
    for row in range(5, ws.max_row+1):
        article = ws.cell(row, 1).value
        if not article:
            empty = True
            new_sizes = {}
            for col in range(4, 16):
                if ws.cell(row, col).value:
                    new_size = str(ws.cell(row, col).value)
                    if new_size in new_sizes:
                        raise Exception(f"У тебя тут размеры повторяются на строке {row}. Как говорит молодеж: ватафак?")
                    new_sizes[col] = new_size
                    empty = False
                else:
                    new_sizes[col] = None
            if not empty:
                sizes = new_sizes
                size_row = row
        else:
            article = str(article)
            box_count_file = ws.cell(row, 2).value
            if not box_count_file:
                raise Exception(f"Какого черта чувак? На строке {row} заполнен артикул, но кол-во коробок пустое. Как это понимать, чел? Исправь, не позорься")
            box_count_file = int(box_count_file)
            models[row] = {
                "article": article,
                "box_count_file": box_count_file,
                "box_count_calc": 0,
                "counts": []
            }
            box_capacity = round((int(ws.cell(row, 3).value) / box_count_file)/5, 0)*5
            if box_capacity == 0:
                box_capacity = 1
            for col in range(4, 16):
                if ws.cell(row, col).value:
                    if sizes[col] == None:
                        raise Exception(f"Короче, у тебя в файле на строке {size_row} определяется к каким размерам относятся какие ячейки, а в ячейке {NUM_TO_COL[col]}{row} есть значение, которое выходит за пределы размерного ряда. Это не дело, чувак. Тут либо надо расширить размерный ряд в строке {size_row}, либо добавить новый размерный ряд между строками {row} и {size_row}. Давай, исправься и попробуем ещё раз. Я верю в тебя.")
                    size_count = int(ws.cell(row, col).value)
                    size_box_count = math.ceil(size_count/box_capacity)
                    models[row]["box_count_calc"] += size_box_count
                    models[row]["counts"].append({"size": sizes[col], "boxes": size_box_count, "pairs": size_count})
    return models

def barcodes_from_ozon(brand):
    try:
        shutil.rmtree(os.path.join(READY_PATH, brand))
    except FileNotFoundError:
        pass
    os.system(f'mkdir "{os.path.join(READY_PATH, brand)}"')
    for model in get_all_data(HEADERS[brand]):
        print(model)
        if model["offer_id"] in ["*/-", "0000 (40)"]:
            continue
        model["offer_id"] = model["offer_id"].replace("81- 2", "81-2")
        model["offer_id"] = model["offer_id"].replace("81- 7", "81-7")
        model["offer_id"] = model["offer_id"].replace(".", "")
        model["offer_id"] = filter_color(model["offer_id"])
        
        article = model["offer_id"].split("(")[0].strip()
        size = model["offer_id"].split("(")[1].strip(") ")
        if article not in os.listdir(os.path.join(READY_PATH, brand)):
            article_path = os.path.join(READY_PATH, brand, article)
            os.makedirs(article_path, exist_ok=True)
        size_path = os.path.join(article_path, size)
        os.makedirs(size_path, exist_ok=True)

        c = Canvas(os.path.join(size_path, f"{size}.pdf"), pagesize=(75 * mm, 120 * mm))
        add_pdf_page(c, brand, model["offer_id"])
        c.save()

def json_to_pdf(brand, models, number):
    for row in models.keys():
        article = models[row]["article"]

        # Создание папки для компании
        brand_path = os.path.join(READY_PATH, brand)
        os.makedirs(brand_path, exist_ok=True)

        # Создание папки для артикула
        article_path = os.path.join(brand_path, article)
        os.makedirs(article_path, exist_ok=True)

        if models[row]["box_count_file"] != models[row]["box_count_calc"]:
            clarify_layout = [
                [sg.Text(models[row]["article"])],
                [],
                [],
                [],
                [sg.OK()]
            ]

            for count in models[row]["counts"]:
                clarify_layout[1].append(sg.Input(count["size"], size=(3, 1), pad=4, background_color=sg.theme_background_color(), text_color="black"))
                clarify_layout[2].append(sg.Input(1, size=(3, 1), pad=4, key="size" + count["size"]))
                clarify_layout[3].append(sg.Input(1, size=(3, 1), pad=4, background_color=sg.theme_background_color(), text_color="black"))

            window = sg.Window("Баркоды", clarify_layout, icon=logo, grab_anywhere=False)
            while True:
                event, values = window.read()
                if event == sg.WIN_CLOSED:
                    break
                if event == "OK":
                    for count in range(len(models[row]["counts"])):  # Обновить значения в counts
                        models[row]["counts"][count]["boxes"] = int(values["size" + models[row]["counts"][count]["size"]])
                    window.close()
                    break

        for count in models[row]["counts"]:
            size = count["size"]

            # Имя PDF-файла для текущего размера
            pdf_file_path = os.path.join(article_path, f"{size}.pdf")  
            c = Canvas(pdf_file_path, pagesize=(75 * mm, 120 * mm))

            for i in range(count["boxes"]):  # Создать страницы в PDF для каждой коробки
                add_pdf_page(c, brand, f"{article} ({size})", number)
            c.save()

def main_menu():
    sg.theme('LightBlue')
    
    layout = [
        [
            sg.Text("Выберите excel файл"),
        ],
        [
            sg.Input(key="excel", expand_x=True), # expand_x для лучшего внешнего вида
            sg.FileBrowse(
                "Выбрать",
                target="excel", # Прямая ссылка на ключ Input
                file_types=(("Лист Microsoft Excel", "*.xlsx"),),
            ),
        ],
        [
            sg.Radio("ARMBEST", group_id="brand", default=True, key="ARMBEST"),
            sg.Radio("BEST 26", group_id="brand", default=False, key="BEST 26")
        ],
        [
            sg.Text("Поставка №"),
            sg.Input(key="number", size=(6, 1))
        ],
        [
            sg.Button("Погнали"),
            sg.Button("Обновить", pad=((230, 5), 3))
        ]
    ]

    window = sg.Window("Баркоды", layout, icon=logo, grab_anywhere=False, finalize=True)

    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:  # if user closes window
            break
        if event == "Погнали":
            if not values["excel"]:
                sg.popup("Не выбран excel", icon=logo)
            elif not values["number"]:
                sg.popup("Введи номер поставки, чел")
            else:
                if values["ARMBEST"]:
                    brand = "ARMBEST"
                else:
                    brand = "BEST 26"
                try:
                    json_to_pdf(brand, excel_to_json(values["excel"]), str(values["number"]))
                except Exception as e:
                    sg.popup(f"Тебе пришла новая ошибка, посмотри вдруг там что-то важное: \n{e}")
        if event == "Обновить":
            try:
                barcodes_from_ozon("ARMBEST")
                barcodes_from_ozon("BEST 26")
            except Exception as e:
                # Исправлен параметр icon (был logo)
                sg.popup(f"Тут ошибка: {e}, кажется у тебя нету доступа к папке с этикетками", icon=logo)

    window.close() # Явное закрытие окна (хорошая практика)

if __name__ == "__main__":
    logo = "black-cock.ico"
    main_menu()