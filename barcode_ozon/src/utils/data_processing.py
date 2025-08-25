# data_processing.py
from flask import json
import requests
import math
from openpyxl import load_workbook

# Переносим константы, связанные с данными
HEADERS = {
    "Armbest": {"Client-Id": "507132", "Api-Key": "token"},
    "BEST 26": {"Client-Id": "2018536", "Api-Key": "token"}
}

NUM_TO_COL = {
    1: "A", 2: "B", 3: "C", 4: "D", 5: "E", 6: "F", 7: "G", 8: "H", 9: "I", 10: "J",
    11: "K", 12: "L", 13: "M", 14: "N", 15: "O", 16: "P", 17: "Q"
}

def filter_color(text):
    # Переносим функцию filter_color как есть
    text = text.replace("бордовые", "brown")
    text = text.replace("бордовая", "brown")
    text = text.replace("бордовый", "brown")
    text = text.replace("бордо", "brown")
    text = text.replace("борд", "brown")
    text = text.replace("бор", "brown")
    text = text.replace("синий", "blue")
    text = text.replace("синие", "blue")
    text = text.replace("синяя", "blue")
    text = text.replace("син", "blue")
    text = text.replace("фиолетовый", "fio")
    text = text.replace("фиолетовые", "fio")
    text = text.replace("фиолетовая", "fio")
    text = text.replace("фиолетов", "fio")
    text = text.replace("фиолето", "fio")
    text = text.replace("фиолет", "fio")
    text = text.replace("фиоле", "fio")
    text = text.replace("фиол", "fio")
    text = text.replace("фио", "fio")
    text = text.replace("фи", "fio")
    text = text.replace("ф", "fio")
    text = text.replace("с", "blue")
    text = text.replace("б", "brown")
    return text

def get_all_data(headers):
    # Переносим функцию get_all_data как есть
    not_first_loop = False
    data = []
    while True:
        if not_first_loop:
            json_data = {"last_id": lastid}
        else:
            json_data = {}
        response = requests.post(
            "https://api-seller.ozon.ru/v2/product/list",
            headers=headers,
            json=json_data,
        )
        response.raise_for_status() # Проверка на ошибки HTTP
        this_data = response.json()
        print(this_data["result"]["total"])
        data += this_data["result"]["items"]
        lastid = this_data["result"]["last_id"]
        if not this_data["result"]["last_id"]:
            break
        not_first_loop = True
        if this_data["result"]["total"] < 1000:
            break
    return data

def load_article_data(json_file_path):
    """
    Загружает данные из JSON файла и создает индекс для быстрого поиска
    
    Args:
        json_file_path (str): Путь к JSON файлу
        
    Returns:
        dict: Словарь для быстрого поиска {артикул: {размер: пары}}
    """
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Создаем индекс для быстрого поиска
        pairs_lookup = {}
        for item in data:
            vendor_code = str(item["Vendorcode"]).strip()
            size = str(item["Size"]).strip()
            pairs = int(item["Pair"])
            
            if vendor_code not in pairs_lookup:
                pairs_lookup[vendor_code] = {}
            pairs_lookup[vendor_code][size] = pairs
            
        return pairs_lookup
    
    except FileNotFoundError:
        raise Exception(f"Файл {json_file_path} не найден")
    except json.JSONDecodeError:
        raise Exception(f"Ошибка чтения JSON файла {json_file_path}")
    except KeyError as e:
        raise Exception(f"В JSON файле отсутствует обязательное поле: {e}")
    except ValueError as e:
        raise Exception(f"Ошибка преобразования данных в JSON файле: {e}")

def load_associations_data(associations_file_path):
    """
    Загружает данные из associations.json и создает индекс для быстрого поиска
    
    Args:
        associations_file_path (str): Путь к associations.json файлу
        
    Returns:
        dict: Словарь для быстрого поиска {full: simple}
    """
    try:
        with open(associations_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Создаем индекс для быстрого поиска: full -> simple
        associations_lookup = {}
        for item in data:
            full = str(item["full"]).strip()
            simple = str(item["simple"]).strip()
            associations_lookup[simple] = full
            
        return associations_lookup
    
    except FileNotFoundError:
        raise Exception(f"Файл {associations_file_path} не найден")
    except json.JSONDecodeError:
        raise Exception(f"Ошибка чтения JSON файла {associations_file_path}")
    except KeyError as e:
        raise Exception(f"В associations.json файле отсутствует обязательное поле: {e}")

def get_pairs_per_box(pairs_lookup, article, size, associations_lookup=None):
    """
    Быстрый поиск количества пар в коробку по артикулу и размеру.
    Если размер не найден, но артикул есть, возвращает 20 пар по умолчанию.
    
    Args:
        pairs_lookup (dict): Индекс из функции load_article_data
        article (str): Артикул
        size (str): Размер
        associations_lookup (dict, optional): Индекс из associations.json
        
    Returns:
        int: Количество пар в коробку
        
    Raises:
        Exception: Если данные не найдены
    """
    article_key = str(article).strip()
    size_key = str(size).strip()
    
    # Попытка 1: Ищем оригинальный артикул с указанным размером
    if article_key in pairs_lookup:
        if size_key in pairs_lookup[article_key]:
            return pairs_lookup[article_key][size_key]
        else:
            # Артикул найден, но размер отсутствует - возвращаем 20 пар по умолчанию
            return 20
    
    # Попытка 2: Используем associations для поиска альтернативного артикула
    if associations_lookup and article_key in associations_lookup:
        simple_article = associations_lookup[article_key]
        if simple_article in pairs_lookup:
            if size_key in pairs_lookup[simple_article]:
                return pairs_lookup[simple_article][size_key]
            else:
                # Альтернативный артикул найден, но размер отсутствует - возвращаем 20 пар по умолчанию
                return 20
    
    # Если артикул не найден нигде
    error_msg = f"Не найден артикул {article_key} в базе данных"
    if associations_lookup and article_key in associations_lookup:
        simple_article = associations_lookup[article_key]
        error_msg += f" (также проверяли артикул {simple_article} из associations)"
    
    raise Exception(error_msg)

def excel_to_json_from_stream(file_stream, json_file_path, association_json_path):
    """
    Основная функция обработки Excel файла с использованием данных из JSON
    
    Args:
        file_stream: Поток Excel файла
        json_file_path (str): Путь к JSON файлу с данными о парах
        
    Returns:
        dict: Результат обработки
    """
    # Загружаем данные из JSON
    pairs_lookup = load_article_data(json_file_path)
    
    wb = load_workbook(file_stream, data_only=True)
    ws = wb.active
    
    sizes = {4:"41", 5:"42", 6:"43", 7:"44", 8:"45", 9:"46", 10:"47", 11:"48", 12:None, 13:None, 14:None, 15:None}
    size_row = 4
    models = {}

    for row in range(5, ws.max_row + 1):
        article = ws.cell(row, 1).value
        if not article:
            empty = True
            new_sizes = {}
            for col in range(4, 16):
                if ws.cell(row, col).value:
                    new_size = str(ws.cell(row, col).value)
                    if new_size in new_sizes:
                        raise Exception(f"У тебя тут размеры повторяются на строке {row}. Как говорит молодеж: ватафак?")
                    new_sizes[col] = new_size
                    empty = False
                else:
                    new_sizes[col] = None
            if not empty:
                sizes = new_sizes
                size_row = row
        else:
            article = str(article)
            box_count_file = ws.cell(row, 2).value
            if not box_count_file:
                raise Exception(f"Какого черта чувак? На строке {row} заполнен артикул, но кол-во коробок пустое.")
            box_count_file = int(box_count_file)
            models[row] = {
                "article": article,
                "box_count_file": box_count_file,
                "box_count_calc": 0,
                "counts": []
            }
            
            # Используем данные из JSON для расчета
            for col in range(4, 16):
                if ws.cell(row, col).value:
                    if sizes[col] is None:
                        raise Exception(f"Размер в ячейке {NUM_TO_COL[col]}{row} выходит за пределы размерного ряда.")
                    
                    size_count = int(ws.cell(row, col).value)
                    
                    # Получаем количество пар в коробку из JSON данных
                    try:
                        pairs_per_box = get_pairs_per_box(pairs_lookup, article, sizes[col], load_associations_data(association_json_path))
                    except Exception as e:
                        raise Exception(f"Ошибка на строке {row}: {str(e)}")
                    
                    # Рассчитываем количество коробок
                    size_box_count = math.ceil(size_count / pairs_per_box)
                    models[row]["box_count_calc"] += size_box_count
                    models[row]["counts"].append({
                        "size": sizes[col],
                        "boxes": size_box_count,
                        "pairs": size_count
                    })
    return models